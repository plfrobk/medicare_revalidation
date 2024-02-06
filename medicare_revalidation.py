from datetime import datetime
from platform import system
from openpyxl import load_workbook
from os import getcwd
from helpers.custom_webdriver import ChromeWebDriver
from helpers.install_chromedriver import InstallChromeDriver

#-----------GLOBAL VARIABLES----------#
isHeadless = True
debug = False
onlyAdjustedDates = False

baseURL = 'https://data.cms.gov/tools/medicare-revalidation-list?size=10&offset=0&npi='
adjustedDatesURL = '&dateFilterType=withAdjustedDueDates'

osRunningScript = system()
if osRunningScript == 'Darwin':
    excelLocation = getcwd() + '/results/Medicare_Revalidation_Results.xlsx'
if osRunningScript == 'Windows':
    excelLocation = getcwd() + '\\results\\Medicare_Revalidation_Results.xlsx'

debugNPIList = ['1184154536', '1386640035', '1053609727'] #Three examples for testing where two should show up with results and one does not

#-----------------MAIN----------------#
try:
    browser = ChromeWebDriver(isHeadless)
except:
    InstallChromeDriver().install_chromedriver()
    browser = ChromeWebDriver(isHeadless)

startTime = datetime.now()
print('Started data pull at: ' + str(startTime))

#-----------------MAIN CODE--------------#
if debug:
    print('Debugging')
    totalRows = len(debugNPIList) + 1
    print('Total manual lookups are: ' + str(totalRows - 1))
else:
    workBook = load_workbook(excelLocation)
    sheet = workBook.active
    totalRows = sheet.max_row
    print('Total rows are: ' + str(totalRows))

for row in range(1,totalRows):
    if debug:
        npiValue = debugNPIList[row-1]
    else:
        rowNumberIncludingHeader = row + 1
        cell = sheet.cell(row=rowNumberIncludingHeader, column=1)
        npiValue = cell.value

    if onlyAdjustedDates:
        browser.get(baseURL + str(npiValue) + adjustedDatesURL)
    else:
        browser.get(baseURL + str(npiValue))
    
    fullResults = browser.grab_element_text('div', 'class', 'ToolsResultsRows', waitBefore=2, timeout=8)
    fullResultsToArray = fullResults.split('\n')

    if fullResults == '':
        organization = '**Not Found**'
        dueDate = '**Not Found**'
        adjustedDueDate = '**Not Found**'
        state = '**Not Found**'
        specialty = '**Not Found**'
        reassignedProviders = '**Not Found**'
        enrollmentType = '**Not Found**' 
    else:
        organization = fullResultsToArray[1]
        dueDate = fullResultsToArray[4]
        adjustedDueDate = 'N/A'
        state = fullResultsToArray[5]
        specialty = fullResultsToArray[6]
        reassignedProviders = fullResultsToArray[7]
        enrollmentType = fullResultsToArray[8]
    
    print('Row ' + str(row) + ' Output - Organization: ' + str(organization) + ', Due Date: ' + str(dueDate) + ', Adjusted Due Date: ' + str(adjustedDueDate) + ', ' + str(state) + ', ' + str(specialty) + ', ' + str(reassignedProviders) + ', ' + str(enrollmentType))

    if debug:
        pass
    else:
        rowName = sheet.cell(row=rowNumberIncludingHeader, column=2)
        rowName.value = str(organization)

        rowDueDate = sheet.cell(row=rowNumberIncludingHeader, column=3)
        try:
            dueDateConverted = datetime.datetime.strptime(dueDate, "%m/%d/%Y")
            rowDueDate.value = dueDateConverted
        except:
            rowDueDate.value = dueDate

        rowAdjustedDueDate = sheet.cell(row=rowNumberIncludingHeader, column=4)
        rowAdjustedDueDate.value = adjustedDueDate

        rowState = sheet.cell(row=rowNumberIncludingHeader, column=5)
        rowState.value = str(state).replace('State: ', '')

        rowSpecialty = sheet.cell(row=rowNumberIncludingHeader, column=6)
        rowSpecialty.value = str(specialty).replace('Specialty: ', '')

        rowReassignedProviders = sheet.cell(row=rowNumberIncludingHeader, column=7)
        try:
            rowReassignedProviders.value = int(str(reassignedProviders).replace('Reassigned Providers: ', ''))
        except:
            rowReassignedProviders.value = reassignedProviders

        rowEnrollmentType = sheet.cell(row=rowNumberIncludingHeader, column=8)
        rowEnrollmentType.value = str(enrollmentType).replace('Enrollment Type: ', '')

        rowLastChecked = sheet.cell(row=rowNumberIncludingHeader, column=9)
        rowLastChecked.value = startTime

if debug:
    pass
else:
    workBook.save(excelLocation)
    browser.close()

endTime = datetime.now()
print('Ended data pull at: ' + str(endTime))